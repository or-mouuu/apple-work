import pandas as pd
from openpyxl import load_workbook
import os

def update_excel_master(excel_path, order_no, date_str, total_quantity):
    """
    Updates the master Excel sheet named 植檢申請流用範本－1.xlsx.
    """
    if not os.path.exists(excel_path):
        return False
        
    try:
        wb = load_workbook(excel_path)
        sheet_name = '植檢番號流用版'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Find the next empty row based on Column A
        empty_row = 1
        for row in range(1, ws.max_row + 2):
            if ws.cell(row=row, column=1).value is None or str(ws.cell(row=row, column=1).value).strip() == '':
                # check if next is also empty to avoid gaps
                if ws.cell(row=row+1, column=1).value is None:
                    empty_row = row
                    break
                    
        # Write Order No to Column A
        ws.cell(row=empty_row, column=1).value = order_no
        # Write Date to Column B
        ws.cell(row=empty_row, column=2).value = date_str
        # Write Total Quantity to somewhere if needed, but original spec just said 注文番號
        
        wb.save(excel_path)
        return True
    except Exception as e:
        raise Exception(f"Failed to update Excel: {str(e)}")
