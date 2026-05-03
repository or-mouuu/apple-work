import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def normalize(s):
    return str(s).replace(" ", "").replace("p", "").replace("up", "").lower()

def get_price(var, grade, size, price_data):
    n_grade = normalize(grade)
    n_var = normalize(var)
    n_size = normalize(size)
    
    target_full = n_var + n_grade
    
    # 1. Strict match on grade and size
    for p in price_data:
        p_var = normalize(p.get('variety', ''))
        if p_var and n_var and p_var != n_var: continue
            
        p_grade = normalize(p.get('grade', ''))
        p_size = normalize(p.get('size', ''))
        
        # Check grade match
        grade_match = False
        p_full = p_var + p_grade
        if p_grade and p_grade in target_full: grade_match = True
        elif target_full in p_full: grade_match = True
        
        if grade_match and p_size == n_size:
            return int(p.get('price', 0))
            
    # 2. Fallback: match only on grade and variety
    for p in price_data:
        p_var = normalize(p.get('variety', ''))
        if p_var and n_var and p_var != n_var: continue
            
        p_grade = normalize(p.get('grade', ''))
        p_full = p_var + p_grade
        
        if p_grade and p_grade in target_full:
            return int(p.get('price', 0))
        elif target_full in p_full:
            return int(p.get('price', 0))
            
    return 0

def create_excel_document(data, price_data, order_no, case_weight, cover_info, output_path, exclude_zero_price=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "INVOICE_PACKINGLIST"
    
    # Font settings
    font_bold = Font(name='Times New Roman', size=10, bold=True)
    font_normal = Font(name='Times New Roman', size=10)
    font_title = Font(name='Times New Roman', size=11, bold=True, italic=True)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_bottom = Border(bottom=thin)
    border_top = Border(top=thin)
    border_tb = Border(top=thin, bottom=thin)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Header Information
    ws['A1'] = "UNIS CO., LTD."
    ws['A1'].font = font_title
    ws['A2'] = "3-4-16 DENEN OAZA HIROSAKI SHI AOMORI KEN JAPAN 036-8086"
    ws['A2'].font = font_normal
    ws['A3'] = "TEL: +81-172-55-8975. FAX: +81-172-55-8976"
    ws['A3'].font = font_normal
    
    ws['D4'] = "INVOICE / PACKINGLIST"
    ws['D4'].font = Font(name='Times New Roman', size=12, bold=True)
    
    # Header block
    ws['A5'] = "SHIPPER :"
    ws['A6'] = cover_info.get("shipper_name", "")
    ws['A7'] = cover_info.get("shipper_addr1", "")
    ws['A8'] = cover_info.get("shipper_addr2", "")
    ws['B9'] = "TEL :  " + cover_info.get("shipper_tel", "")
    ws['B10'] = "FAX :  " + cover_info.get("shipper_fax", "")
    
    ws['D5'] = "REF.NO."
    ws['E5'] = order_no
    ws['H5'] = "DATE"
    ws['I5'] = cover_info.get("date", "")
    
    ws['D7'] = "BOOKING AGENT"
    ws['E7'] = cover_info.get("booking_agent", "")
    ws['H7'] = "BOOKING NO."
    ws['I7'] = cover_info.get("booking_no", "")
    
    ws['D8'] = "SHIPPED PER MV"
    ws['F8'] = cover_info.get("shipped_per", "")
    
    ws['D9'] = "FROM :"
    ws['E9'] = cover_info.get("from_port", "")
    ws['G9'] = "TO :"
    ws['H9'] = cover_info.get("to_port", "")
    
    ws['D10'] = "ON OR ABOUT"
    ws['F10'] = cover_info.get("on_or_about", "")
    
    ws['A11'] = "CONSIGNEE:"
    ws['A12'] = cover_info.get("consignee_name", "")
    ws['A13'] = cover_info.get("consignee_addr1", "")
    ws['A14'] = cover_info.get("consignee_addr2", "")
    ws['B15'] = "TEL :  " + cover_info.get("consignee_tel", "")
    ws['B16'] = "FAX :  " + cover_info.get("consignee_fax", "")
    
    ws['D11'] = "REMARKS :"
    ws['E12'] = "ORIGIN: " + cover_info.get("origin", "")
    ws['E14'] = "BRAND: " + cover_info.get("brand", "")
    ws['E15'] = "PALLET: " + str(cover_info.get("pallet", ""))
    ws['G15'] = "ICE BOX"
    
    ws['A18'] = "NOTIFY PATY:"
    ws['D18'] = "ALSO NOTIFY :"
    
    # Apply fonts and borders for header block
    for row in range(5, 20):
        for col in range(1, 10):
            ws.cell(row=row, column=col).font = font_normal
            
    # Table Headers
    row_idx = 21
    headers = ["MARKS AND NOS", "DESCRIPTION", "", "QUANTITY", "", "UNIT PRICE", "AMOUNT", "NET", "GROSS"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = header
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = border_tb
        
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="FRESH APPLE").font = font_normal
    ws.cell(row=row_idx, column=6, value="C&F: keelung").font = font_normal
    ws.cell(row=row_idx, column=6).alignment = align_center
    
    row_idx += 1
    sub_headers = ["NO MARK", "", "KG", "", "CASE", "", "", "KG", "KG"]
    for col_idx, sh in enumerate(sub_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = sh
        cell.font = font_normal
        cell.alignment = align_center
        
    row_idx += 2
    
    # Data Processing
    last_var = None
    last_grade = None
    
    total_case = 0
    total_amount = 0
    total_net = 0
    total_gross = 0
    
    for item in data:
        var = str(item.get('variety', '')).strip()
        grade = str(item.get('grade', '')).strip()
        size = str(item.get('size', ''))
        qty = int(item.get('quantity', 0))
        
        price = get_price(var, grade, size, price_data)
        
        if exclude_zero_price and price == 0:
            continue
            
        amount = qty * price
        net = qty * case_weight
        gross = qty * (case_weight + 1.0)
        
        total_case += qty
        total_amount += amount
        total_net += net
        total_gross += gross
        
        is_new_var = var != last_var
        is_new_grade = is_new_var or (grade != last_grade)
        
        # Space between different groups
        if is_new_grade and last_grade is not None:
            row_idx += 1
            
        if is_new_var:
            ws.cell(row=row_idx, column=1, value=var).font = font_normal
            
        if is_new_grade:
            ws.cell(row=row_idx, column=2, value=grade).font = font_normal
            
        ws.cell(row=row_idx, column=3, value=f"{size} p").font = font_normal
        ws.cell(row=row_idx, column=3).alignment = align_right
        
        ws.cell(row=row_idx, column=4, value=case_weight).font = font_normal
        ws.cell(row=row_idx, column=4).alignment = align_center
        
        ws.cell(row=row_idx, column=5, value=qty).font = font_normal
        ws.cell(row=row_idx, column=5).alignment = align_center
        
        ws.cell(row=row_idx, column=6, value=f"¥{price:,.0f}").font = font_normal
        ws.cell(row=row_idx, column=6).alignment = align_right
        
        ws.cell(row=row_idx, column=7, value=f"¥{amount:,.0f}").font = font_normal
        ws.cell(row=row_idx, column=7).alignment = align_right
        
        ws.cell(row=row_idx, column=8, value=f"{net:.1f}").font = font_normal
        ws.cell(row=row_idx, column=8).alignment = align_right
        
        ws.cell(row=row_idx, column=9, value=f"{gross:.2f}").font = font_normal
        ws.cell(row=row_idx, column=9).alignment = align_right
        
        last_var = var
        last_grade = grade
        row_idx += 1
        
    row_idx += 1
    
    # Bottom Total Border
    for col in range(1, 10):
        ws.cell(row=row_idx, column=col).border = border_top
        ws.cell(row=row_idx, column=col).font = font_bold
        
    ws.cell(row=row_idx, column=2, value="TOTAL")
    ws.cell(row=row_idx, column=5, value=f"{total_case:,}")
    ws.cell(row=row_idx, column=5).alignment = align_center
    ws.cell(row=row_idx, column=7, value=f"¥{total_amount:,.0f}")
    ws.cell(row=row_idx, column=7).alignment = align_right
    ws.cell(row=row_idx, column=8, value=f"{total_net:,.1f}")
    ws.cell(row=row_idx, column=8).alignment = align_right
    ws.cell(row=row_idx, column=9, value=f"{total_gross:,.0f}")
    ws.cell(row=row_idx, column=9).alignment = align_right
    
    row_idx += 4
    ws.cell(row=row_idx, column=6, value="UNIS CO.,LTD.").font = font_title
    
    # Save the workbook
    wb.save(output_path)
    return output_path
